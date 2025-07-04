from django.contrib import messages
from django.shortcuts import render, redirect
from openpyxl.reader.excel import load_workbook
from rest_framework import viewsets, serializers
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db import transaction

from inventoryTracker import settings
from inventoryTracker.inventory.forms import ProductImportForm
from inventoryTracker.inventory.models import Product, Category, Manufacturer, Warehouse, Shelf, Vendor
from inventoryTracker.inventory.serializers import ProductSerializer, CategorySerializer, ManufacturerSerializer, \
    WareHouseSerializer, ShelfSerializer, VendorSerializer

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from decimal import Decimal
from datetime import datetime


def index(request):
    return render(request, 'index.html', {
        'api_url_base': settings.API_BASE_URL
    })


@login_required
def products_view(request):
    context = {'api_url_base': settings.API_BASE_URL}
    return render(request, 'products.html', context)


@login_required
def warehouse_view(request):
    context = {'api_url_base': settings.API_BASE_URL}
    return render(request, 'warehouses.html', context)


@login_required
def shelves_view(request):
    context = {'api_url_base': settings.API_BASE_URL}
    return render(request, 'shelf.html', context)


@login_required
def vendors_view(request):
    context = {'api_url_base': settings.API_BASE_URL}
    return render(request, 'vendors.html', context)


@login_required
def manufacturer_view(request):
    context = {'api_url_base': settings.API_BASE_URL}
    return render(request, 'manufacturers.html', context)


@login_required
def category_view(request):
    context = {'api_url_base': settings.API_BASE_URL}
    return render(request, 'categories.html', context)


class WareHouseViewSet(viewsets.ModelViewSet):
    serializer_class = WareHouseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Warehouse.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ShelfViewSet(viewsets.ModelViewSet):
    serializer_class = ShelfSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Shelf.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        warehouse_id = self.request.data.get('warehouse')
        try:
            warehouse = Warehouse.objects.get(id=warehouse_id, user=self.request.user)
            serializer.save(user=self.request.user, warehouse=warehouse)
        except Warehouse.DoesNotExist:
            raise serializers.ValidationError({'warehouse': 'Invalid warehouse selection'})

    def perform_update(self, serializer):
        warehouse_id = self.request.data.get('warehouse')
        try:
            warehouse = Warehouse.objects.get(id=warehouse_id, user=self.request.user)
            serializer.save(warehouse=warehouse)
        except Warehouse.DoesNotExist:
            raise serializers.ValidationError({'warehouse': 'Invalid warehouse selection'})


class VendorViewSet(viewsets.ModelViewSet):
    serializer_class = VendorSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Vendor.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ManufacturerViewSet(viewsets.ModelViewSet):
    serializer_class = ManufacturerSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Manufacturer.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class CategoryViewSet(viewsets.ModelViewSet):
    serializer_class = CategorySerializer

    def get_queryset(self):
        return Category.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]
    queryset = Product.objects.none()

    def get_queryset(self):
        return Product.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        vendor_ids = self.request.data.get('vendor', [])
        product = serializer.save(user=self.request.user)

        # Convert vendor IDs to integers if they're strings
        if vendor_ids:
            try:
                vendor_ids = [int(vid) for vid in vendor_ids]
                product.vendor.set(vendor_ids)
            except (ValueError, TypeError) as e:
                raise serializers.ValidationError({'vendor': 'Invalid vendor IDs'})

    def perform_update(self, serializer):
        vendor_ids = self.request.data.get('vendor', [])
        product = serializer.save()

        if 'vendor' in self.request.data:
            try:
                vendor_ids = [int(vid) for vid in vendor_ids]
                product.vendor.set(vendor_ids)
            except (ValueError, TypeError) as e:
                raise serializers.ValidationError({'vendor': 'Invalid vendor IDs'})

    @action(detail=False, methods=['post'])
    def bulk_assign_category(self, request):
        product_ids = request.data.get('product_ids', [])
        category_id = request.data.get('category_id')

        if not product_ids or not category_id:
            return Response({'error': 'product_ids and category_id are required'}, status=400)

        try:
            category = Category.objects.get(id=category_id, user=request.user)
            products = Product.objects.filter(id__in=product_ids, user=request.user)
            updated_count = products.update(category=category)
            return Response({'success': f'Updated {updated_count} products'})
        except Category.DoesNotExist:
            return Response({'error': 'Category not found or not owned by user'}, status=404)

    @action(detail=False, methods=['post'])
    def bulk_remove_category(self, request):
        product_ids = request.data.get('product_ids', [])

        if not product_ids:
            return Response({'error': 'product_ids are required'}, status=400)

        products = Product.objects.filter(id__in=product_ids, user=request.user)
        updated_count = products.update(category=None)
        return Response({'success': f'Removed category from {updated_count} products'})

    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        product_ids = request.data.get('product_ids', [])

        if not product_ids:
            return Response({'error': 'product_ids are required'}, status=400)

        products = Product.objects.filter(id__in=product_ids, user=request.user)
        count = products.count()
        products.delete()
        return Response({'success': f'Deleted {count} products'})

    @action(detail=False, methods=['post'])
    def bulk_assign_warehouse(self, request):
        product_ids = request.data.get('product_ids', [])
        warehouse_id = request.data.get('warehouse_id')

        if not product_ids or not warehouse_id:
            return Response({'error': 'product_ids and warehouse_id are required'}, status=400)

        try:
            warehouse = Warehouse.objects.get(id=warehouse_id, user=request.user)
            products = Product.objects.filter(id__in=product_ids, user=request.user)
            updated_count = products.update(warehouse=warehouse)
            return Response({'success': f'Updated {updated_count} products'})
        except Warehouse.DoesNotExist:
            return Response({'error': 'Warehouse not found or not owned by user'}, status=404)

    @action(detail=False, methods=['post'])
    def bulk_remove_warehouse(self, request):
        product_ids = request.data.get('product_ids', [])

        if not product_ids:
            return Response({'error': 'product_ids are required'}, status=400)

        products = Product.objects.filter(id__in=product_ids, user=request.user)
        updated_count = products.update(warehouse=None)
        return Response({'success': f'Removed warehouse from {updated_count} products'})

    @action(detail=False, methods=['post'])
    def bulk_assign_shelf(self, request):
        product_ids = request.data.get('product_ids', [])
        shelf_id = request.data.get('shelf_id')

        if not product_ids or not shelf_id:
            return Response({'error': 'product_ids and shelf_id are required'}, status=400)

        try:
            shelf = Shelf.objects.get(id=shelf_id, user=request.user)
            products = Product.objects.filter(id__in=product_ids, user=request.user)
            updated_count = products.update(shelf=shelf)
            return Response({'success': f'Updated {updated_count} products'})
        except Shelf.DoesNotExist:
            return Response({'error': 'Shelf not found or not owned by user'}, status=404)

    @action(detail=False, methods=['post'])
    def bulk_remove_shelf(self, request):
        product_ids = request.data.get('product_ids', [])

        if not product_ids:
            return Response({'error': 'product_ids are required'}, status=400)

        products = Product.objects.filter(id__in=product_ids, user=request.user)
        updated_count = products.update(shelf=None)
        return Response({'success': f'Removed shelf from {updated_count} products'})

    @action(detail=False, methods=['post'])
    def bulk_assign_manufacturer(self, request):
        product_ids = request.data.get('product_ids', [])
        manufacturer_id = request.data.get('manufacturer_id')

        if not product_ids or not manufacturer_id:
            return Response({'error': 'product_ids and manufacturer_id are required'}, status=400)

        try:
            manufacturer = Manufacturer.objects.get(id=manufacturer_id, user=request.user)
            products = Product.objects.filter(id__in=product_ids, user=request.user)
            updated_count = products.update(manufacturer=manufacturer)
            return Response({'success': f'Updated {updated_count} products'})
        except Manufacturer.DoesNotExist:
            return Response({'error': 'Manufacturer not found or not owned by user'}, status=404)

    @action(detail=False, methods=['post'])
    def bulk_remove_manufacturer(self, request):
        product_ids = request.data.get('product_ids', [])

        if not product_ids:
            return Response({'error': 'product_ids are required'}, status=400)

        products = Product.objects.filter(id__in=product_ids, user=request.user)
        updated_count = products.update(manufacturer=None)
        return Response({'success': f'Removed manufacturer from {updated_count} products'})

    @action(detail=False, methods=['post'])
    def bulk_assign_vendor(self, request):
        product_ids = request.data.get('product_ids', [])
        vendor_id = request.data.get('vendor_id')

        if not product_ids or not vendor_id:
            return Response({'error': 'product_ids and vendor_id are required'}, status=400)

        try:
            products = Product.objects.filter(id__in=product_ids, user=request.user)
            vendor = Vendor.objects.get(id=vendor_id, user=request.user)

            for product in products:
                product.vendor.add(vendor)

            return Response({'success': f'Added vendor to {len(products)} products'})
        except Vendor.DoesNotExist:
            return Response({'error': 'Vendor not found or not owned by user'}, status=404)

    @action(detail=False, methods=['post'])
    def bulk_remove_vendor(self, request):
        product_ids = request.data.get('product_ids', [])
        vendor_id = request.data.get('vendor_id')

        if not product_ids or not vendor_id:
            return Response({'error': 'product_ids and vendor_id are required'}, status=400)

        try:
            products = Product.objects.filter(id__in=product_ids, user=request.user)
            vendor = Vendor.objects.get(id=vendor_id, user=request.user)

            for product in products:
                product.vendor.remove(vendor)

            return Response({'success': f'Removed vendor from {len(products)} products'})
        except Vendor.DoesNotExist:
            return Response({'error': 'Vendor not found or not owned by user'}, status=404)


def export_products_to_excel(request):
    # Get products for the current user
    products = Product.objects.filter(user=request.user)

    # Create a workbook and add worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Create header row with styling
    headers = [
        "Name", "Category", "SKU", "Barcode", "Manufacturer",
        "Model", "Model 2", "Model 3", "Model 4", "Model 5",
        "Quantity", "Warehouse", "Shelf", "Box", "Bag",
        "Vendors", "Buy Price", "Sell Price", "Created At", "Updated At", "Additional Info"
    ]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws[f"{col_letter}1"].font = Font(bold=True)

    # Add data rows
    for row_num, product in enumerate(products, 2):
        ws[f"A{row_num}"] = product.name
        ws[f"B{row_num}"] = product.category.name if product.category else ""
        ws[f"C{row_num}"] = product.sku
        ws[f"D{row_num}"] = product.barcode
        ws[f"E{row_num}"] = product.manufacturer.name if product.manufacturer else ""
        ws[f"F{row_num}"] = product.model
        ws[f"G{row_num}"] = product.model_2
        ws[f"H{row_num}"] = product.model_3
        ws[f"I{row_num}"] = product.model_4
        ws[f"J{row_num}"] = product.model_5
        ws[f"K{row_num}"] = product.quantity
        ws[f"L{row_num}"] = product.warehouse.name if product.warehouse else ""
        ws[f"M{row_num}"] = product.shelf.name if product.shelf else ""
        ws[f"N{row_num}"] = product.box
        ws[f"O{row_num}"] = product.bag
        ws[f"P{row_num}"] = ", ".join([v.name for v in product.vendor.all()])
        ws[f"Q{row_num}"] = float(product.buy_price) if product.buy_price else 0
        ws[f"R{row_num}"] = float(product.sell_price) if product.sell_price else 0
        ws[f"S{row_num}"] = product.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ws[f"T{row_num}"] = product.updated_at.strftime("%Y-%m-%d %H:%M:%S")
        ws[f"U{row_num}"] = product.additional_info

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products_export.xlsx'
    wb.save(response)

    return response


def import_products(request):
    if request.method == 'POST':
        form = ProductImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    file = request.FILES['file']
                    wb = load_workbook(filename=file, data_only=True)
                    ws = wb.active

                    # Expected headers mapping
                    expected_headers = {
                        'Name': 'name',
                        'Category': 'category',
                        'SKU': 'sku',
                        'Barcode': 'barcode',
                        'Manufacturer': 'manufacturer',
                        'Model': 'model',
                        'Model 2': 'model_2',
                        'Model 3': 'model_3',
                        'Model 4': 'model_4',
                        'Model 5': 'model_5',
                        'Quantity': 'quantity',
                        'Warehouse': 'warehouse',
                        'Shelf': 'shelf',
                        'Box': 'box',
                        'Bag': 'bag',
                        'Vendors': 'vendor',
                        'Buy Price': 'buy_price',
                        'Sell Price': 'sell_price',
                        'Additional Info': 'additional_info'
                    }

                    # Validate headers
                    headers = [cell.value for cell in ws[1]]
                    missing_headers = set(expected_headers.keys()) - set(headers)
                    if missing_headers:
                        raise ValueError(f"Missing required columns: {', '.join(missing_headers)}")

                    # Create column mapping
                    column_map = {header: idx + 1 for idx, header in enumerate(headers)}

                    success_count = 0
                    error_count = 0
                    errors = []

                    for row in ws.iter_rows(min_row=2):  # Skip header row
                        row_num = row[0].row
                        try:
                            row_data = {'user': request.user}

                            for header, field in expected_headers.items():
                                if header in ['Vendors', 'Category', 'Manufacturer', 'Warehouse', 'Shelf']:
                                    continue  # Handle these separately

                                cell_value = ws.cell(row=row_num, column=column_map[header]).value
                                if cell_value is None:
                                    continue

                                if header in ['Buy Price', 'Sell Price']:
                                    row_data[field] = Decimal(str(cell_value))
                                elif header == 'Quantity':
                                    row_data[field] = int(cell_value)
                                else:
                                    row_data[field] = str(cell_value).strip()

                            if 'Category' in column_map:
                                category_name = ws.cell(row=row_num, column=column_map['Category']).value
                                if category_name:
                                    if form.cleaned_data['create_missing']:
                                        category, _ = Category.objects.get_or_create(
                                            name=str(category_name).strip(),
                                            user=request.user
                                        )
                                        row_data['category'] = category
                                    else:
                                        try:
                                            row_data['category'] = Category.objects.get(
                                                name=str(category_name).strip(),
                                                user=request.user
                                            )
                                        except Category.DoesNotExist:
                                            raise ValueError(f"Category '{category_name}' does not exist")

                            if 'Manufacturer' in column_map:
                                manufacturer_name = ws.cell(row=row_num, column=column_map['Manufacturer']).value
                                if manufacturer_name:
                                    if form.cleaned_data['create_missing']:
                                        manufacturer, _ = Manufacturer.objects.get_or_create(
                                            name=str(manufacturer_name).strip(),
                                            user=request.user
                                        )
                                        row_data['manufacturer'] = manufacturer
                                    else:
                                        try:
                                            row_data['manufacturer'] = Manufacturer.objects.get(
                                                name=str(manufacturer_name).strip(),
                                                user=request.user
                                            )
                                        except Manufacturer.DoesNotExist:
                                            raise ValueError(f"Manufacturer '{manufacturer_name}' does not exist")

                            if 'Warehouse' in column_map:
                                warehouse_name = ws.cell(row=row_num, column=column_map['Warehouse']).value
                                if warehouse_name:
                                    if form.cleaned_data['create_missing']:
                                        warehouse, _ = Warehouse.objects.get_or_create(
                                            name=str(warehouse_name).strip(),
                                            user=request.user
                                        )
                                        row_data['warehouse'] = warehouse
                                    else:
                                        try:
                                            row_data['warehouse'] = Warehouse.objects.get(
                                                name=str(warehouse_name).strip(),
                                                user=request.user
                                            )
                                        except Warehouse.DoesNotExist:
                                            raise ValueError(f"Warehouse '{warehouse_name}' does not exist")

                            if 'Shelf' in column_map:
                                shelf_name = ws.cell(row=row_num, column=column_map['Shelf']).value
                                if shelf_name:
                                    warehouse = row_data.get('warehouse')
                                    if not warehouse:
                                        raise ValueError("Shelf requires a warehouse to be specified")

                                    if form.cleaned_data['create_missing']:
                                        shelf, _ = Shelf.objects.get_or_create(
                                            name=str(shelf_name).strip(),
                                            warehouse=warehouse,
                                            user=request.user
                                        )
                                        row_data['shelf'] = shelf
                                    else:
                                        try:
                                            row_data['shelf'] = Shelf.objects.get(
                                                name=str(shelf_name).strip(),
                                                warehouse=warehouse,
                                                user=request.user
                                            )
                                        except Shelf.DoesNotExist:
                                            raise ValueError(
                                                f"Shelf '{shelf_name}' does not exist in warehouse '{warehouse.name}'")

                            vendor_names = []
                            if 'Vendors' in column_map:
                                vendor_cell = ws.cell(row=row_num, column=column_map['Vendors']).value
                                if vendor_cell:
                                    vendor_names = [
                                        name.strip()
                                        for name in str(vendor_cell).split(',')
                                        if name.strip()
                                    ]

                            # Check if product exists (if update_existing)
                            product = None
                            if form.cleaned_data['update_existing'] and 'sku' in row_data:
                                product = Product.objects.filter(
                                    sku=row_data['sku'],
                                    user=request.user
                                ).first()

                            if product:
                                # Update existing product
                                for field, value in row_data.items():
                                    if field != 'user' and field != 'sku':
                                        setattr(product, field, value)
                                product.full_clean()
                                product.save()
                            else:
                                # Create new product
                                product = Product(**row_data)
                                product.full_clean()
                                product.save()

                            if vendor_names:
                                vendors = []
                                for name in vendor_names:
                                    if form.cleaned_data['create_missing']:
                                        vendor, _ = Vendor.objects.get_or_create(
                                            name=name,
                                            user=request.user
                                        )
                                    else:
                                        try:
                                            vendor = Vendor.objects.get(
                                                name=name,
                                                user=request.user
                                            )
                                        except Vendor.DoesNotExist:
                                            raise ValueError(f"Vendor '{name}' does not exist")
                                    vendors.append(vendor)
                                product.vendor.set(vendors)

                            success_count += 1

                        except Exception as e:
                            error_count += 1
                            errors.append(f"Row {row_num}: {str(e)}")
                            continue

                    messages.success(request, f"Successfully processed {success_count} products")
                    if error_count > 0:
                        messages.warning(request, f"Failed to process {error_count} rows")
                        request.session['import_errors'] = errors[:50]  # Limit to 50 errors

                    return redirect('products')

            except Exception as e:
                messages.error(request, f"Import failed: {str(e)}")
    else:
        form = ProductImportForm()

    import_errors = request.session.pop('import_errors', None)

    return render(request, 'import.html', {
        'form': form,
        'import_errors': import_errors
    })
